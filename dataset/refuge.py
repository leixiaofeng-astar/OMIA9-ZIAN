from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import logging
import os
from collections import OrderedDict
from scipy.io import loadmat, savemat
from openpyxl import load_workbook
import csv
import cv2
import numpy as np
from sklearn.model_selection import train_test_split
import random
from dataset.dataset import FoveaDataset


logger = logging.getLogger(__name__)

def data_split(full_list, ratio=0.8, shuffle=True, random_status=1200):
    """
    数据集拆分: 将列表full_list按比例ratio（随机）划分为2个子列表sublist_1与sublist_2
    :param full_list: 数据列表
    :param ratio:     子列表1
    :param shuffle:   子列表2
    :return:
    """
    n_total = len(full_list)
    offset = int(n_total * ratio)
    if n_total == 0 or offset < 1:
        return [], full_list
    if shuffle:
        random.Random(random_status).shuffle(full_list)
    sublist_1 = full_list[:offset]
    sublist_2 = full_list[offset:]
    return sublist_1, sublist_2


def crop_fundus_image(img_data):
    h, w, c = img_data.shape
    start_h = int(h // 3)
    end_h = int(h // 3 * 2)
    start_w = int(w // 3)
    end_w = int(w // 3 * 2)
    crop_img = img_data[start_h:end_h, start_w:end_w, :]
    return crop_img


class Dataset(FoveaDataset):
    def __init__(self, cfg, root, image_set, is_train, transform=None):
        super(Dataset, self).__init__(cfg, root, image_set, is_train, transform)
        self.cfg = cfg
        self.image_set = image_set
        self.db = self._get_db(image_set)

        # test_img = cv2.imread(self.db[0]['image'], cv2.IMREAD_COLOR)
        # test_img = self.db[0]['image']
        # self.db_image_size = np.array([test_img.shape[1], test_img.shape[0]])

        # each image in dataset could have different image size
        self.db_image_size = []
        db_len = len(self.db)
        for i in range(db_len):
            test_img = self.db[i]['image']
            self.db_image_size.append(np.array([test_img.shape[1], test_img.shape[0]]))  # H, W, C

        if is_train and cfg.DATASET.TRAIN_FOLD > 0:
            np.random.seed(1234) # fix seed
            perm = np.random.permutation(len(self.db))
            perm = np.split(perm, 5, axis=0)
            del perm[cfg.DATASET.TRAIN_FOLD - 1] # remove this fold
            indices = np.concatenate(perm, axis=0)
            db = [self.db[_] for _ in indices]
            self.db = db

        logger.info('=> load {} samples'.format(len(self.db)))

    def is_image_file(self, img_path):
        img_name = os.path.basename(img_path)
        _, ext = os.path.splitext(img_name)
        if ext == '.jpg' or ext == '.jpeg' or ext == '.png' or ext == '.JPG' or ext == '.JPEG':
            return True
        else:
            return False

    def _get_db(self, image_set):
        '''
        ├── AGE-test
        ├── AGE-test-GT
        ├── AGE-train
        ├── AGE-train-GT
        ├── GAMMA
        ├── refuge1-test
        ├── refuge1-test-GT
        ├── refuge1-train
        └── refuge1-train-GT
        '''
        TRAIN_AGE_MODEL = self.cfg.TRAIN.AGE_DATA
        TRAIN_GAMMA_MODEL = self.cfg.TRAIN.GAMMA_DATA

        if TRAIN_AGE_MODEL:
            # AGE SS localization challenge
            print("Dataset preparation for AGE model on fovea localization")
            if image_set == 'train':
                # training images and labels
                train_anno_filename = os.path.join(self.root, 'AGE-train-GT', 'Train_Fovea_locations_AGE.xlsx')

                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                train_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero
                    fx = float(booksheet.cell(row=i, column=3).value) - 1
                    fy = float(booksheet.cell(row=i, column=4).value) - 1
                    fname = booksheet.cell(row=i, column=2).value
                    image_file = os.path.join(self.root, 'AGE-train', fname)

                    if (not self.is_image_file(image_file)) or (not os.path.isfile(image_file)):
                        continue

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)

                    # note: some image has wrong GT, pick out them from database
                    if fx > 0 and fy > 0:
                        train_db.append({
                            'image': data_numpy,
                            'fovea': np.array([fx, fy], np.float32),
                            'filename': image_file,
                        })

                train_db, _ = data_split(train_db, ratio=0.8, shuffle=True, random_status=1200)
                return train_db
            elif image_set == 'test':
                # test images and labels
                test_anno_filename = os.path.join(self.root, 'AGE-test-GT', 'test_Fovea_locations_AGE.xlsx')

                workbook = load_workbook(test_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                test_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero
                    fx = float(booksheet.cell(row=i, column=4).value) - 1
                    fy = float(booksheet.cell(row=i, column=5).value) - 1
                    fname = booksheet.cell(row=i, column=2).value
                    image_file = os.path.join(self.root, 'AGE-test', fname)
                    if not self.is_image_file(image_file): continue

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                    test_db.append({
                        'image': data_numpy,
                        'fovea': np.array([fx, fy], np.float32),
                        'filename': image_file,
                    })
                return test_db
            elif image_set == 'val':
                # training images and labels
                train_anno_filename = os.path.join(self.root, 'AGE-train-GT', 'Train_Fovea_locations_AGE.xlsx')
                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                val_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero
                    fx = float(booksheet.cell(row=i, column=3).value) - 1
                    fy = float(booksheet.cell(row=i, column=4).value) - 1
                    fname = booksheet.cell(row=i, column=2).value
                    image_file = os.path.join(self.root, 'AGE-train', fname)

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                    val_db.append({
                        'image': data_numpy,
                        'fovea': np.array([fx, fy], np.float32),
                        'filename': image_file,
                    })
                _, val_db, = data_split(val_db, ratio=0.8, shuffle=True, random_status=1200)
                return val_db
            else:
                assert('Unknown image set: %s' % (image_set))

        elif TRAIN_GAMMA_MODEL:
            # TODO
            # GAMMA fovea localization challenge
            print("Dataset preparation for GAMMA fovea localization ({})".format(image_set))
            if image_set == 'train' or image_set == 'val':
                # training images and labels
                train_anno_filename = os.path.join(self.root, 'GAMMA', 'fovea_localization_training_GT.xlsx')

                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                train_db = []
                pickup_list = []

                My_list = [*range(2, booksheet.max_row + 1, 1)]
                index_train, index_test = train_test_split(My_list, test_size=0.15, random_state=4)

                if image_set == 'train':
                    pickup_list = index_train
                elif image_set == 'val':
                    pickup_list = index_test

                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero
                    fx = float(booksheet.cell(row=i, column=2).value) - 1
                    fy = float(booksheet.cell(row=i, column=3).value) - 1
                    fname = booksheet.cell(row=i, column=1).value

                    # TODO
                    if i in pickup_list:
                        fname = fname + ".jpg"
                        # print("filename: {}".format(fname))
                        image_file = os.path.join(self.root, 'GAMMA', 'train', fname)

                        if (not self.is_image_file(image_file)) or (not os.path.isfile(image_file)):
                            continue

                        data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)

                        # note: some image has wrong GT, pick out them from database
                        if fx > 0 and fy > 0:
                            train_db.append({
                                'image': data_numpy,
                                'fovea': np.array([fx, fy], np.float32),
                                'filename': image_file,
                            })
                return train_db
            if image_set == 'train+val':
                # training images and labels
                train_anno_filename = os.path.join(self.root, 'GAMMA', 'fovea_localization_training_GT.xlsx')

                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                train_db = []
                pickup_list = []

                pickup_list = [*range(2, booksheet.max_row + 1, 1)]

                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero
                    fx = float(booksheet.cell(row=i, column=2).value) - 1
                    fy = float(booksheet.cell(row=i, column=3).value) - 1
                    fname = booksheet.cell(row=i, column=1).value

                    # TODO
                    if i in pickup_list:
                        fname = fname + ".jpg"
                        # print("filename: {}".format(fname))
                        image_file = os.path.join(self.root, 'GAMMA', 'train', fname)

                        if (not self.is_image_file(image_file)) or (not os.path.isfile(image_file)):
                            continue

                        data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)

                        # note: some image has wrong GT, pick out them from database
                        if fx > 0 and fy > 0:
                            train_db.append({
                                'image': data_numpy,
                                'fovea': np.array([fx, fy], np.float32),
                                'filename': image_file,
                            })
                return train_db
            if image_set == 'val2':
                # GAMMA validation dataset
                # TODO
                gamma_final_challenge = True
                # training images and labels
                if not gamma_final_challenge:
                    train_anno_filename = os.path.join(self.root, 'GAMMA', 'fovea_localization_val_GT.xlsx')
                else:
                    train_anno_filename = os.path.join(self.root, 'GAMMA', 'fovea_localization_test_GT.xlsx')

                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                val_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero
                    fx = float(booksheet.cell(row=i, column=2).value) - 1
                    fy = float(booksheet.cell(row=i, column=3).value) - 1

                    fname = booksheet.cell(row=i, column=1).value
                    fname = fname + '.jpg'
                    if not gamma_final_challenge:
                        image_file = os.path.join(self.root, 'GAMMA', 'val', fname)
                    else:
                        image_file = os.path.join(self.root, 'GAMMA', 'test', fname)

                    if (not self.is_image_file(image_file)) or (not os.path.isfile(image_file)):
                        continue

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                    # note: some image has wrong GT, pick out them from database
                    if fx > 0 and fy > 0:
                        val_db.append({
                            'image': data_numpy,
                            'fovea': np.array([fx, fy], np.float32),
                            'filename': image_file,
                        })
                return val_db

            elif image_set == 'test':
                # test images and labels
                raise("Wait for GAMMA data release!!!")

            # elif image_set == 'val':
                # TODO: validation images and labels
                
                return val_db
            else:
                assert ('Unknown image set: %s' % (image_set))

        else:
            # NOTE: *************  REFUGE  *************
            # Refuge fovea localization challenge
            print("Dataset preparation for Refuge model on fovea localization")

            if image_set == 'train+val':
                # training images and labels
                # train_anno_filename = os.path.join(self.root, 'refuge2-train-GT', 'refuge_Fovea_location.xlsx')
                train_anno_filename = os.path.join(self.root, 'refuge1-train-GT', 'refuge1_train_Fovea_location.xlsx')
                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                train_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero (0, 0)
                    fx = float(booksheet.cell(row=i, column=3).value) - 1
                    fy = float(booksheet.cell(row=i, column=4).value) - 1
                    fname = booksheet.cell(row=i, column=2).value
                    # image_file = os.path.join(self.root, 'refuge2-train', fname)
                    image_file = os.path.join(self.root, 'refuge1-train', fname)

                    if not self.is_image_file(image_file):
                        print("{} file cannot be found!!!" .format(image_file))
                        continue

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                    train_db.append({
                        'image': data_numpy,
                        'fovea': np.array([fx, fy], np.float32),
                        'filename': image_file,
                    })

                return train_db

            elif image_set == 'train':
                # training images and labels
                # train_anno_filename = os.path.join(self.root, 'refuge2-train-GT', 'refuge_Fovea_location.xlsx')
                train_anno_filename = os.path.join(self.root, 'refuge1-train-GT', 'refuge1_train_Fovea_location.xlsx')
                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                train_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero (0, 0)
                    fx = float(booksheet.cell(row=i, column=3).value) - 1
                    fy = float(booksheet.cell(row=i, column=4).value) - 1
                    fname = booksheet.cell(row=i, column=2).value
                    # image_file = os.path.join(self.root, 'refuge2-train', fname)
                    image_file = os.path.join(self.root, 'refuge1-train', fname)

                    if not self.is_image_file(image_file):
                        print("{} file cannot be found!!!".format(image_file))
                        continue

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                    train_db.append({
                        'image': data_numpy,
                        'fovea': np.array([fx, fy], np.float32),
                        'filename': image_file,
                    })

                train_db, _ = data_split(train_db, ratio=0.8, shuffle=True, random_status=1200)
                return train_db

            elif image_set == 'val':
                # training images and labels
                # train_anno_filename = os.path.join(self.root, 'refuge2-train-GT', 'refuge_Fovea_location.xlsx')
                train_anno_filename = os.path.join(self.root, 'refuge1-train-GT', 'refuge1_train_Fovea_location.xlsx')
                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                train_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero (0, 0)
                    fx = float(booksheet.cell(row=i, column=3).value) - 1
                    fy = float(booksheet.cell(row=i, column=4).value) - 1
                    fname = booksheet.cell(row=i, column=2).value
                    # image_file = os.path.join(self.root, 'refuge2-train', fname)
                    image_file = os.path.join(self.root, 'refuge1-train', fname)

                    if not self.is_image_file(image_file):
                        print("{} file cannot be found!!!".format(image_file))
                        continue

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                    train_db.append({
                        'image': data_numpy,
                        'fovea': np.array([fx, fy], np.float32),
                        'filename': image_file,
                    })

                _, train_db, = data_split(train_db, ratio=0.8, shuffle=True, random_status=1200)
                return train_db

            elif image_set == 'test':
                # test images and labels
                # training images and labels
                # train_anno_filename = os.path.join(self.root, 'refuge2-test-GT', 'Fovea_locations_dummy_refuge2_val.xlsx')
                train_anno_filename = os.path.join(self.root, 'refuge1-test-GT', 'refuge1_test_Fovea_location.xlsx')
                workbook = load_workbook(train_anno_filename)
                booksheet = workbook.active
                rows = booksheet.rows
                columns = booksheet.columns
                train_db = []
                for i, row in enumerate(rows, 1):
                    if i == 1: continue  # skip the first row
                    # substract 1 pixel as we assume indexing from zero (0, 0)
                    # fx = float(booksheet.cell(row=i, column=4).value) - 1
                    # fy = float(booksheet.cell(row=i, column=5).value) - 1
                    fx = float(booksheet.cell(row=i, column=3).value) - 1
                    fy = float(booksheet.cell(row=i, column=4).value) - 1
                    fname = booksheet.cell(row=i, column=2).value
                    # image_file = os.path.join(self.root, 'refuge2-test', fname)
                    image_file = os.path.join(self.root, 'refuge1-test', fname)

                    if not self.is_image_file(image_file):
                        print("{} file cannot be found!!!".format(image_file))
                        continue

                    data_numpy = cv2.imread(image_file, cv2.IMREAD_COLOR)
                    train_db.append({
                        'image': data_numpy,
                        'fovea': np.array([fx, fy], np.float32),
                        'filename': image_file,
                    })

                return train_db

            elif image_set == 'val2':
                # TODO: refuge2 validation db -- xiaofeng comment
                raise("Warning: we don't support it now!!!")
            else:
                assert ('Unknown image set: %s' % (image_set))


    def evaluate(self, preds, output_dir, debug_enable=False):
        num_images = len(self.db)
        assert num_images == len(self.db)

        # the predicted coordinates are based on the resized and center-cropped
        # images, convert them back
        image_size = self.cfg.MODEL.IMAGE_SIZE  # resized image size
        crop_size = self.cfg.MODEL.CROP_SIZE
        # xiaofeng test for AGE x128 size issue
        # crop_size = self.cfg.MODEL.PATCH_SIZE
        # end of xiaofeng change for AGE x128 size issue

        pw = (image_size[0] - crop_size[0]) // 2
        ph = (image_size[1] - crop_size[1]) // 2
        preds[:, 0] += pw
        preds[:, 1] += ph

        for x in range(num_images):
            preds[x, 0] *= (self.db_image_size[x][0] * 1.0 / image_size[0])
            preds[x, 1] *= (self.db_image_size[x][1] * 1.0 / image_size[1])

        l2_dist_sum = 0.
        sdr_5px_cnt = 0
        sdr_10px_cnt = 0
        sdr_15px_cnt = 0
        sdr_20px_cnt = 0
        for _ in range(num_images):
            gt = self.db[_]['fovea']
            # print("file:{}, fovea:[{},{}]/{}" .format(self.db[_]['filename'], preds[_,0],preds[_,1], self.db[_]['fovea']))
            l2_per_img = np.sqrt(np.sum((preds[_, :] - gt)**2))

            # add for SDR
            if l2_per_img < 20:
                sdr_20px_cnt += 1
                if l2_per_img < 15:
                    sdr_15px_cnt += 1
                    if l2_per_img < 10:
                        sdr_10px_cnt += 1
                        if l2_per_img < 5:
                            sdr_5px_cnt += 1

            l2_dist_sum += l2_per_img
        l2_dist_avg = l2_dist_sum / num_images

        if output_dir is not None:
            csv_file = os.path.join(output_dir, 'fovea_location_results.csv')
            with open(csv_file, 'w') as f:
                cw = csv.writer(f, delimiter=",", lineterminator="\n")
                if debug_enable:
                    cw.writerow(['ImageName', 'Fovea_X', 'Fovea_Y', 'FoveaX_GT', 'FoveaY_GT', 'X_bias', 'Y_bias'])
                    for _ in range(num_images):
                        image_name = os.path.basename(self.db[_]['filename'])
                        # image_name = str(os.path.splitext(image_name)[0])
                        fovea_x = '%.2f' % (preds[_, 0]+1)
                        fovea_y = '%.2f' % (preds[_, 1]+1)
                        FoveaX_GT = '%.2f' % (self.db[_]['fovea'][0])
                        FoveaY_GT = '%.2f' % (self.db[_]['fovea'][1])
                        X_bias = '%.2f' % (preds[_, 0] - self.db[_]['fovea'][0])
                        Y_bias = '%.2f' % (preds[_, 1] - self.db[_]['fovea'][1])
                        cw.writerow([image_name, fovea_x, fovea_y, FoveaX_GT, FoveaY_GT, X_bias, Y_bias])
                else:
                    cw.writerow(['ImageName', 'Fovea_X', 'Fovea_Y'])
                    for _ in range(num_images):
                        image_name = os.path.basename(self.db[_]['filename'])
                        # image_name = str(os.path.splitext(image_name)[0])
                        # TODO : check if +1 if necessary
                        fovea_x = '%.2f' %(preds[_, 0] + 1)
                        fovea_y = '%.2f' %(preds[_, 1] + 1)
                        cw.writerow([image_name, fovea_x, fovea_y])
                f.close()

        # add SDR
        sdr_5px = sdr_5px_cnt/num_images
        sdr_10px = sdr_10px_cnt/num_images
        sdr_15px = sdr_15px_cnt/num_images
        sdr_20px = sdr_20px_cnt/num_images
        logger.info('SDR 5px:{}; 10px:{}; 15px: {}; 20px: {}; L2:{}'.format(sdr_5px, sdr_10px, sdr_15px, sdr_20px, l2_dist_avg))
        return l2_dist_avg
