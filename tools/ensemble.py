"""

This module is to ensemble different model for refuge challenge

Example:
    Examples can be given using either the ``Example`` or ``Examples``
    sections. Sections support any reStructuredText formatting, including
    literal blocks::
    Run on GPU to convert torch GPU train model to CPU status
        $ python3 tools/ensemble.py --cfg experiments/refuge.yaml \
            --prevModelDir output/refuge/fovea_net/refuge \
            --logDir ../final-submission \
            --release True

        copy the csv file to logDir, then combine them togother,
        then use Density-Based Spatial Clustering to get the cluster number
        then lock the maximum cluster and calculate the average

"""
from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import argparse
import os
import pprint
import importlib
from sklearn.cluster import DBSCAN

import torch
import torch.nn.parallel
import torch.backends.cudnn as cudnn
import torch.optim
import torch.utils.data
import torch.utils.data.distributed
import torchvision.transforms as transforms

import _init_paths
from config.common import _C as cfg
from config.common import update_config
from core.loss import HybridLoss
from core.function import validate
from utils.utils import create_logger
import numpy as np
import pandas as pd
import time
import glob
from collections import Counter
from openpyxl import load_workbook

import dataset
import models


def parse_args():
    parser = argparse.ArgumentParser(description='Test keypoints network')
    # general
    parser.add_argument('--cfg',
                        help='experiment configure file name',
                        required=True,
                        type=str)

    parser.add_argument('opts',
                        help="Modify config options using the command-line",
                        default=None,
                        nargs=argparse.REMAINDER)

    parser.add_argument('--modelDir',
                        help='model directory',
                        type=str,
                        default='')
    parser.add_argument('--logDir',
                        help='log directory',
                        type=str,
                        default='')
    parser.add_argument('--dataDir',
                        help='data directory',
                        type=str,
                        default='')
    parser.add_argument('--prevModelDir',
                        help='prev Model directory',
                        type=str,
                        default='')
    parser.add_argument('--release',
                        help='select different dataset',
                        type=bool, default=False)

    args = parser.parse_args()
    return args


def main():
    args = parse_args()

    model_folder = args.prevModelDir
    if model_folder is not None:
        if not os.path.isdir(model_folder):
            raise ("No existing model_folder!!!")

    outputs_dir = args.logDir
    if not os.path.exists(outputs_dir):
        os.makedirs(outputs_dir)

    update_config(cfg, args)

    logger, final_output_dir, tb_log_dir = create_logger(
        cfg, args.cfg, 'valid')

    logger.info(pprint.pformat(args))
    logger.info(cfg)

    # cudnn related setting
    cudnn.benchmark = cfg.CUDNN.BENCHMARK
    torch.backends.cudnn.deterministic = cfg.CUDNN.DETERMINISTIC
    torch.backends.cudnn.enabled = cfg.CUDNN.ENABLED

    #  model file / run command / prediction result folder
    # TEST.MODEL_FILE model_file
    csv_default_file = "Localization_Results.csv"
    csv_location_dict = {'DS_FINAL': './',
                         'HR_FINAL': './output/refuge/fovea_net/refuge',
                         # 'P2_DS': '../best_batch4_version/',
                         # 'P2_ROI': '../best_batch4_version/log',
                         # 'P2_FINAL': '../best_batch4_version/output/refuge/fovea_net/refuge',
                         # 'HR_DS': '../test_batch4/',
                         # 'HR_ROI': '../test_batch4/log',
                         # 'HR_FINAL': '../test_batch4/output/refuge/fovea_net/refuge',
                         }
    release_option = args.release
    print("release option: {}" .format(release_option))
    if release_option is False:
        cmd_opt_1 = 'TEST.RELEASE_TEST False'
    else:
        cmd_opt_1 = 'TEST.RELEASE_TEST True'

    models_list = [

        # 1
        # ['model_best_L1037.pth',  # ED=0.00638
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TEST.TRIAL_RUN True TEST.TRIAL_DATASET 3 TRAIN.DATA_CLAHE True PRINT_FREQ 100',
        #  'HR_FINAL'],

        # 2
        # ['model_best_L990_train_LR4e4_E100.pth', # ED=0.00668
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
        #  'HR_FINAL'],

        # 3
        # ['model_best_L977.pth',   # ED=0.00629 TL=10.14
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100', 'HR_FINAL'],

        # 4
        ['model_best_L987_LR2e4_e100_be73.pth',   # 5th ED=0.00579 TL=10.30
         'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
         'HR_FINAL'],

        # 5
        ['model_best_L943_LR2e4.pth',  # ED=0.00541 TL=9.84
         'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
         'HR_FINAL'],

        # 6
        # ['model_best_L871_LR2e4_e100_be43.pth',  # 6th ED=0.0062
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
        #  'HR_FINAL'],

        # # 7
        # ['model_best_HRNet_L891_TL1164_LR4e5.pth',  # ED=0.0063 TL=11.64
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
        #  'HR_FINAL'],
        #
        # # 8
        # ['model_best_ResUnet_E140_LR1e6.pth',  # ED=0.00863  TL=19.09
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True MODEL.HRNET_TYPE 2 PRINT_FREQ 100',
        #  'HR_FINAL'],
        #
        # # 9
        # ['model_best_UNET_L778_TL1429_LR2r5_E110.pth',  # ED=xxx  TL=14.29
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True MODEL.HRNET_TYPE 1 PRINT_FREQ 100',
        #  'HR_FINAL'],

        # 10
        # ['model_best_HRNET_Final_E200_LR4e5_TL1468_E140.pth',  # ED=0.00692 TL=11.64
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
        #  'HR_FINAL'],

        # 11
        # ['model_best_ResUnet_Final_LR1e5_TL1303_E100.pth', # ED=0.00659  TL=13.03  'model_best_ResUnet_Final_E140_TL1699.pth',  # ED=xxx  TL=16.99
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True MODEL.HRNET_TYPE 2 PRINT_FREQ 100',
        #  'HR_FINAL'],

        # 12
        # ['model_best_UNET_full_LR2r5_TL15_E80.pth',  # ED=xxx  TL=15.76
        #  'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True MODEL.HRNET_TYPE 1 PRINT_FREQ 100',
        #  'HR_FINAL'],

        # 13
        ['model_best_HRNET_Final_Retrain_LR1e5_TL1060_E40.pth',  # ED=0.00515 TL=10.60
         'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
         'HR_FINAL'],

        # 14
        ['model_best_HRNET_Final_E120_LR6e6_TL1030_E40.pth',  # ED=0.00548 TL=10.30
         'python3 tools/test.py --cfg experiments/refuge.yaml TRAIN.GAMMA_DATA True MODEL.TRIP_ROI True MODEL.SELF_ATTEN True TEST.RELEASE_TEST True TRAIN.DATA_CLAHE True PRINT_FREQ 100',
         'HR_FINAL'],


        # ensemble model 1: 3-9(7 models): ED=0.00575 TL=9.84
        # ensemble model 2: 4-6, 10-12(6 models): ED=0.0074 TL=9.61
        # ensemble model 3: 4-5, 13-14(4 models): ED=0.00531 TL=10.33
    ]

    model_file_list = []
    for model_file in glob.glob(os.path.join(model_folder, '*.pth')):
        model_file_list.append(model_file)
        print("Found model file: ", model_file)

    fovea_location_groups = []
    model_idx = 0
    for model_file in model_file_list:
        model_found = False
        model_name = "Model"
        delete_existing_csv(csv_location_dict, csv_default_file)
        for record in models_list:
            if record[0] in model_file:
                model_name = str(os.path.splitext(record[0])[0])
                if 'P2' in record[0]:
                    cmd_str = record[1] + ' ' + 'TEST.MODEL_FILE' + ' ' + model_file
                else:
                    cmd_str = record[1] + ' ' + 'TEST.MODEL_FILE' + ' ' + model_file + ' ' + cmd_opt_1
                print("execute command: %s" %(cmd_str))

                # TODO
                os.system(cmd_str)
                time.sleep(1)
                os.system("sync")
                print()

                csv_path = csv_location_dict[record[2]]
                csv_access_path = os.path.join(csv_path, csv_default_file)
                model_idx += 1
                model_found = True
                break

        if model_found:
            # Read the csv localization info
            print("read_csv: %s" % (csv_access_path))
            try:
                excel_data_df = pd.read_csv(csv_access_path)
            except:
                time.sleep(1)
                print("read_csv again: %s" % (csv_access_path))
                os.system("sync")
                time.sleep(2)
                excel_data_df = pd.read_csv(csv_access_path)

            # Fovea_X_column = '%s_Fovea_X'%(model_name)
            # Fovea_Y_column = '%s_Fovea_Y' % (model_name)
            Fovea_X_column = 'Model_%d_Fovea_X' % (model_idx)
            Fovea_Y_column = 'Model_%d_Fovea_Y' % (model_idx)
            excel_data_df.rename(columns={'Fovea_X': Fovea_X_column, 'Fovea_Y': Fovea_Y_column}, inplace=True)
            # excel_data_df['Final_Fovea_X'] = None
            # excel_data_df['Final_Fovea_Y'] = None
            fovea_location_groups.append(excel_data_df)

    # Get the final csv
    master_csv_file = 'master_fovea.csv'
    master_csv_fullname = os.path.join(outputs_dir, master_csv_file)

    # TODO
    if len(fovea_location_groups) > 0:
        with open(master_csv_fullname, 'w') as f:
            df = pd.concat(fovea_location_groups, axis=1)
            df = df.loc[:, ~df.columns.duplicated()]
            # df=df.T.drop_duplicates().T
            df.to_csv(f, index=False)
            # pd.merge(fovea_location_groups, on=['data', 'Final_Fovea_X', 'Final_Fovea_Y'], how='right').to_csv(f, ignore_index=True)


    # Apply Density-Based Spatial Clustering
    excel_data_df = pd.read_csv(master_csv_fullname)
    model_sum = model_idx
    location_lists = []
    filename_list = excel_data_df['data'].tolist()
    record_sum = len(filename_list)
    print("Ensemble {} model, record number in csv is {}" .format(model_sum, record_sum))

    Final_Fovea_X_list = []
    Final_Fovea_Y_list = []
    # Final_Fovea_X_list = excel_data_df['Final_Fovea_X']
    # Final_Fovea_Y_list = excel_data_df['Final_Fovea_Y']
    # import pdb
    # pdb.set_trace()
    for i in range(1, model_sum+1):
        Fovea_X_column = 'Model_%d_Fovea_X' % (i)
        Fovea_Y_column = 'Model_%d_Fovea_Y' % (i)
        location_lists.append(excel_data_df[Fovea_X_column].tolist())
        location_lists.append(excel_data_df[Fovea_Y_column].tolist())

    # calculate the L2 for validation
    model_in_csv = (len(location_lists)+1)//2
    pred_list = []
    for i in range(record_sum):
        image_fovea_lists = []
        try:
            for j in range(model_in_csv):
                image_fovea_lists.append([location_lists[j*2][i], location_lists[j*2+1][i]])
        except:
            import pdb
            pdb.set_trace()

        arr = np.array(image_fovea_lists)
        # print("Image {} Fovea prediction: {};  Result: {}" .format(i, arr.shape, arr))

        eps_init = 5
        avg_pos = get_max_cluster_average(arr, eps_init, more_than_half=False, radius_step=2)
        # print("output final fovea: ", avg_pos)

        Final_Fovea_X_list.append(avg_pos[0])
        Final_Fovea_Y_list.append(avg_pos[1])
        pred_list.append([avg_pos[0], avg_pos[1]])

    excel_data_df['Fovea_X'] = Final_Fovea_X_list
    excel_data_df['Fovea_Y'] = Final_Fovea_Y_list

    master_csv_file = 'submit_master_fovea.csv'
    master_csv_fullname = os.path.join(outputs_dir, master_csv_file)
    excel_data_df.to_csv(master_csv_fullname, index=False)

    pred_array = np.array(pred_list)
    print("prediction array shape: ", pred_array.shape)
    evaluate_L2_distance(pred_array, release_option)


def delete_existing_csv(path_dict, csv_filename):
    for key in path_dict.keys():
        path = path_dict[key]
        csv_file = os.path.join(path, csv_filename)
        cmd_str = 'rm -f ' + csv_file
        print("execute command: %s" % (cmd_str))
        os.system(cmd_str)
        os.system("sync")


def evaluate_L2_distance(preds, release_option=True):
    if release_option:
        annotation_filename = os.path.join('../../miccai-refuge/data/', 'GAMMA', 'fovea_localization_val_GT.xlsx')
    else:
        annotation_filename = os.path.join('../../miccai-refuge/data/', 'GAMMA', 'fovea_localization_training_GT.xlsx')
    print("Ground truth filename: ", annotation_filename)
    workbook = load_workbook(annotation_filename)
    booksheet = workbook.active
    rows = booksheet.rows
    columns = booksheet.columns
    val2_db = []
    num_images = 0
    for i, row in enumerate(rows, 1):
        if i == 1: continue  # skip the first row
        # substract 1 pixel as we assume indexing from zero
        fx = float(booksheet.cell(row=i, column=2).value) - 1
        fy = float(booksheet.cell(row=i, column=3).value) - 1
        fname = booksheet.cell(row=i, column=1).value
        val2_db.append({
            'fovea': np.array([fx, fy], np.float32),
            'data': fname,
        })
        num_images += 1

    l2_dist_sum = 0.0
    for _ in range(num_images):
        gt = val2_db[_]['fovea']
        item_dist = np.sqrt(np.sum((preds[_, :] - gt) ** 2))
        print('Item[{}] L2 Distance: {} -- pred: {} vs gt: {}' .format(_, item_dist, preds[_, :], gt))
        l2_dist_sum += item_dist
    l2_dist_avg = l2_dist_sum / num_images
    print('Average L2 Distance on test set: l2_dist_avg = %.2f (total_images=%d)' % (l2_dist_avg, num_images))
    return l2_dist_avg


class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count if self.count != 0 else 0

def get_max_cluster_average(np_array, radius, more_than_half=True, radius_step=3):
    X = np_array
    eps_init = radius
    flag_continoue = True
    # print("input format: ", X.shape)

    while flag_continoue:
        clustering = DBSCAN(eps=eps_init, min_samples=1).fit(X)
        cluster_labels = clustering.labels_.tolist()
        # print(cluster_labels)
        maxlabel = max(cluster_labels, key=cluster_labels.count)
        # print("maxlabel index:", maxlabel)
        # cluster_labels = [0, 1, 1, 0, 2, 3, 4]
        result = Counter(cluster_labels)
        # print("Radius is %d, Matched number: %d" %(eps_init, result[maxlabel]))
        if more_than_half:
            min_num = X.shape[0] // 2 + 1
        else:
            min_num = X.shape[0] // 2
        if result[maxlabel] >= min_num:
            flag_continoue = False
            final_dists = AverageMeter()
            for i in range(X.shape[0]):
                if cluster_labels[i] == maxlabel:
                    final_dists.update(X[i])
                    # print("Index {} has {}" .format(i, X[i]))
            # print("avg: ", final_dists.avg)
        else:
            eps_init += radius_step

    return np.rint(final_dists.avg)


def test1():
    X = np.array([[800, 800], [1068, 1021], [1070, 1050],\
                  [810, 840], [1075, 1049], [1020, 1000], [1021, 1036]])
    print(X.shape)
    eps_init = 20
    avg_pos = get_max_cluster_average(X, eps_init, more_than_half=False)
    print(avg_pos)


if __name__ == '__main__':
    main()
    print("GAMMA Fovea Ensemble Program Exit ... \n")
