"""

This module is to ensemble different model for refuge challenge

Example:
    Examples can be given using either the ``Example`` or ``Examples``
    sections. Sections support any reStructuredText formatting, including
    literal blocks::
    Run on GPU to convert torch GPU train model to CPU status
        $ python3 tools/calc_csv_fovea.py --logDir ../final-submission/csv_table --release True

        get the csv file to logDir, then combine them togother,
        then use Density-Based Spatial Clustering to get the cluster number
        then lock the maximum cluster and calculate the average

"""
from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import argparse
import os
import pprint
import importlib
from sklearn.cluster import DBSCAN

import torch
import torch.nn.parallel
import torch.backends.cudnn as cudnn
import torch.optim
import torch.utils.data
import torch.utils.data.distributed
import torchvision.transforms as transforms


import numpy as np
import pandas as pd
import time
import glob
from collections import Counter
from openpyxl import load_workbook



def parse_args():
    parser = argparse.ArgumentParser(description='Test keypoints network')
    # general
    parser.add_argument('--cfg',
                        help='experiment configure file name',
                        required=False,
                        type=str)

    parser.add_argument('opts',
                        help="Modify config options using the command-line",
                        default=None,
                        nargs=argparse.REMAINDER)

    parser.add_argument('--modelDir',
                        help='model directory',
                        type=str,
                        default='')
    parser.add_argument('--logDir',
                        help='log directory',
                        type=str,
                        default='')
    parser.add_argument('--dataDir',
                        help='data directory',
                        type=str,
                        default='')
    parser.add_argument('--prevModelDir',
                        help='prev Model directory',
                        type=str,
                        default='')
    parser.add_argument('--model_num',
                        help='total model number',
                        type=int, default=7)
    parser.add_argument('--release',
                        help='select different dataset',
                        type=bool, default=False)

    args = parser.parse_args()
    return args


def main():
    args = parse_args()

    outputs_dir = args.logDir
    if not os.path.exists(outputs_dir):
        os.makedirs(outputs_dir)

    model_num = args.model_num
    release_option = args.release

    # Read all the csv localization info
    master_csv_file = 'master_fovea.csv'
    master_csv_fullname = os.path.join(outputs_dir, master_csv_file)
    print("read_csv: %s" % (master_csv_fullname))

    # Apply Density-Based Spatial Clustering
    excel_data_df = pd.read_csv(master_csv_fullname)
    model_sum = model_num
    location_lists = []
    filename_list = excel_data_df['ImageName'].tolist()
    record_sum = len(filename_list)
    print("The total record number in csv is ", record_sum)

    Final_Fovea_X_list = []
    Final_Fovea_Y_list = []
    # Final_Fovea_X_list = excel_data_df['Final_Fovea_X']
    # Final_Fovea_Y_list = excel_data_df['Final_Fovea_Y']
    # import pdb
    # pdb.set_trace()
    for i in range(1, model_sum+1):
        Fovea_X_column = 'Model_%d_Fovea_X' % (i)
        Fovea_Y_column = 'Model_%d_Fovea_Y' % (i)
        location_lists.append(excel_data_df[Fovea_X_column].tolist())
        location_lists.append(excel_data_df[Fovea_Y_column].tolist())

    # calculate the L2 for validation
    model_in_csv = (len(location_lists)+1)//2
    pred_list = []
    for i in range(record_sum):
        image_fovea_lists = []
        try:
            for j in range(model_in_csv):
                image_fovea_lists.append([location_lists[j*2][i], location_lists[j*2+1][i]])
        except:
            import pdb
            pdb.set_trace()

        arr = np.array(image_fovea_lists)
        # print("Image {} Fovea prediction: {};  Result: {}" .format(i, arr.shape, arr))
        eps_init = 8
        avg_pos = get_max_cluster_average(arr, eps_init, more_than_half=True)
        print("[{}] output final fovea: {}" .format(i, avg_pos))

        Final_Fovea_X_list.append(avg_pos[0])
        Final_Fovea_Y_list.append(avg_pos[1])
        pred_list.append([avg_pos[0], avg_pos[1]])

    excel_data_df['Fovea_X'] = Final_Fovea_X_list
    excel_data_df['Fovea_Y'] = Final_Fovea_Y_list

    master_csv_file = 'converted_master_fovea.csv'
    master_csv_fullname = os.path.join(outputs_dir, master_csv_file)
    excel_data_df.to_csv(master_csv_fullname, index=False)
    print("Final submission file: ", master_csv_fullname)

    pred_array = np.array(pred_list)
    print("prediction array shape: ", pred_array.shape)
    evaluate_L2_distance(pred_array, release_option)


def evaluate_L2_distance(preds, release_option=False):
    if not release_option:
        annotation_filename = os.path.join('../data', 'REFUGE-Test-GT', 'Glaucoma_label_and_Fovea_location.xlsx')
    else:
        annotation_filename = os.path.join('../data', 'Refuge2-Validation-GT', 'Fovea_locations_dummy.xlsx')
    print("Ground truth filename: ", annotation_filename)
    workbook = load_workbook(annotation_filename)
    booksheet = workbook.active
    rows = booksheet.rows
    columns = booksheet.columns
    val2_db = []
    num_images = 0
    for i, row in enumerate(rows, 1):
        if i == 1: continue  # skip the first row
        # substract 1 pixel as we assume indexing from zero
        fx = float(booksheet.cell(row=i, column=4).value) - 1
        fy = float(booksheet.cell(row=i, column=5).value) - 1
        fname = booksheet.cell(row=i, column=2).value
        val2_db.append({
            'fovea': np.array([fx, fy], np.float32),
            'filename': fname,
        })
        num_images += 1

    l2_dist_sum = 0.0
    for _ in range(num_images):
        gt = val2_db[_]['fovea']
        item_dist = np.sqrt(np.sum((preds[_, :] - gt) ** 2))
        print('Item[{}] L2 Distance: {} -- pred: {} vs gt: {}' .format(_, item_dist, preds[_, :], gt))
        l2_dist_sum += item_dist
    l2_dist_avg = l2_dist_sum / num_images
    print('Average L2 Distance on test set: l2_dist_avg = %.2f (num_images=%d)' % (l2_dist_avg, num_images))
    return l2_dist_avg


class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count if self.count != 0 else 0

def get_max_cluster_average(np_array, radius, more_than_half=True, radius_step=5):
    X = np_array
    eps_init = radius
    flag_continoue = True
    # print("input format: ", X.shape)

    while flag_continoue:
        clustering = DBSCAN(eps=eps_init, min_samples=1).fit(X)
        cluster_labels = clustering.labels_.tolist()
        # print(cluster_labels)
        maxlabel = max(cluster_labels, key=cluster_labels.count)
        # print("maxlabel index:", maxlabel)
        # cluster_labels = [0, 1, 1, 0, 2, 3, 4]
        result = Counter(cluster_labels)
        print("Radius is %d, Matched number: %d" %(eps_init, result[maxlabel]))
        if more_than_half:
            min_num = X.shape[0] // 2 + 1
        else:
            min_num = X.shape[0] // 2
        if result[maxlabel] >= min_num:
            flag_continoue = False
            final_dists = AverageMeter()
            for i in range(X.shape[0]):
                if cluster_labels[i] == maxlabel:
                    final_dists.update(X[i])
                    # print("Index {} has {}" .format(i, X[i]))
            # print("avg: ", final_dists.avg)
        else:
            eps_init += radius_step

    return np.rint(final_dists.avg)


def test1():
    X = np.array([[800, 800], [1068, 1021], [1070, 1050],\
                  [810, 840], [1075, 1049], [1020, 1000], [1021, 1036]])
    print(X.shape)
    eps_init = 20
    avg_pos = get_max_cluster_average(X, eps_init, more_than_half=False)
    print(avg_pos)


if __name__ == '__main__':
    main()
    print("Calculate Refuge Fovea Ensemble Program Exit ... \n")
